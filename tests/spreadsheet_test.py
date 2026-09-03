#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Tests for the spreadsheet bridge (v0.66.0 — Adoption Roadmap Goal 7).

Covers:
- CSV export of the three tabular views (headers, rows, quoting)
- XLSX export (openpyxl; skipped when the extra is not installed)
- value import (CSV/XLSX) into evaluator bindings, incl. units
- CLI: `view --format csv`, `eval --set-file`, `sysmlpy xlsx`
- end-to-end: spreadsheet values gate constraints
"""

import csv
import io
import sys

import pytest

from sysmlpy import loads, check_constraints, evaluate_expression
from sysmlpy.spreadsheet import (
    tabular_view_to_csv, data_value_tabular_to_csv,
    relationship_matrix_to_csv, write_xlsx,
    import_values_csv, import_values_xlsx, parse_value_literal,
)

MODEL = """package VehicleSpec {
    part def Vehicle {
        attribute mass : Real := 1200;
        attribute speed : Real := 25.0;
        part wheels: Wheel[4];
        constraint c1 { mass > 1000 }
        constraint c3 { mass > 5000 }
    }
    part def Wheel { attribute radius : Real := 0.3 [m]; }
}"""


def model():
    return loads(MODEL)


# ---------------------------------------------------------------------------
# parse_value_literal
# ---------------------------------------------------------------------------


class TestParseValueLiteral:

    def test_int(self):
        assert parse_value_literal("42") == 42

    def test_float(self):
        assert parse_value_literal("25.0") == 25.0

    def test_bool(self):
        assert parse_value_literal("true") is True
        assert parse_value_literal("false") is False

    def test_unit_string(self):
        from sysmlpy.usage import ureg
        assert parse_value_literal("1200 kg") == 1200 * ureg.kg
        assert parse_value_literal("80 km/h") == 80 * ureg("km/h")

    def test_string_fallback(self):
        assert parse_value_literal("VehicleSpec") == "VehicleSpec"


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------


class TestCsvExport:

    def test_tabular_header_and_rows(self):
        text = tabular_view_to_csv(loads(MODEL))
        rows = list(csv.reader(io.StringIO(text)))
        assert rows[0] == ["Name", "Type", "Kind", "Parent",
                           "Typed By", "Specializes"]
        names = {r[0] for r in rows[1:]}
        assert {"Vehicle", "Wheel", "mass", "radius", "wheels"} <= names

    def test_tabular_csv_via_view_format(self):
        from sysmlpy.plantuml import as_tabular_view
        direct = as_tabular_view(loads(MODEL), output_format="csv")
        assert list(csv.reader(io.StringIO(direct)))[0][0] == "Name"

    def test_data_value_csv(self):
        text = data_value_tabular_to_csv(loads(MODEL))
        rows = list(csv.reader(io.StringIO(text)))
        assert rows[0] == ["Element", "Attribute", "Value", "Unit", "Type"]
        body = {tuple(r) for r in rows[1:]}
        assert ("Wheel", "radius", "0.3", "meter", "Real") in body

    def test_matrix_csv(self):
        text = relationship_matrix_to_csv(loads(MODEL))
        rows = list(csv.reader(io.StringIO(text)))
        assert len(rows) >= 2
        # first column holds row labels (element names)
        assert rows[0][0] == "Vehicle"

    def test_csv_quoting(self):
        m = loads('package P { part def V { attribute note : String '
                  ':= "a, b"; } }')
        text = data_value_tabular_to_csv(m)
        assert '"a, b"' in text  # embedded comma is quoted

    def test_focus_changes_output(self):
        m = loads(MODEL)
        full = tabular_view_to_csv(m)
        focused = tabular_view_to_csv(m, focus="Vehicle")
        assert focused != full
        # focus lists only the focused subtree — the sibling def is gone
        assert "Wheel" not in focused


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------

def _has_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


needs_openpyxl = pytest.mark.skipif(
    not _has_openpyxl(), reason="openpyxl not installed"
)


class TestXlsxExport:

    @needs_openpyxl
    def test_workbook_has_all_sheets(self, tmp_path):
        out = tmp_path / "model.xlsx"
        write_xlsx(loads(MODEL), out)
        import openpyxl
        wb = openpyxl.load_workbook(out)
        assert set(wb.sheetnames) == {"Tabular", "DataValues", "Matrix"}

    @needs_openpyxl
    def test_sheet_selection(self, tmp_path):
        out = tmp_path / "sel.xlsx"
        write_xlsx(loads(MODEL), out, include=("data_value",))
        import openpyxl
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["DataValues"]

    @needs_openpyxl
    def test_unknown_sheet_kind_raises(self, tmp_path):
        with pytest.raises(ValueError):
            write_xlsx(loads(MODEL), tmp_path / "x.xlsx",
                       include=("bogus",))

    @needs_openpyxl
    def test_header_row_bold(self, tmp_path):
        out = tmp_path / "b.xlsx"
        write_xlsx(loads(MODEL), out, include=("tabular",))
        import openpyxl
        ws = openpyxl.load_workbook(out)["Tabular"]
        assert all(c.font.bold for c in ws[1])

    def test_missing_openpyxl_error(self, monkeypatch, tmp_path):
        # simulate ImportError from the import machinery
        orig_import = __import__

        def fake_import(name, *a, **kw):
            if name == "openpyxl":
                raise ImportError("no openpyxl")
            return orig_import(name, *a, **kw)

        monkeypatch.setattr("builtins.__import__", fake_import)
        with pytest.raises(ImportError, match="openpyxl"):
            write_xlsx(loads(MODEL), tmp_path / "x.xlsx")


# ---------------------------------------------------------------------------
# value import
# ---------------------------------------------------------------------------


class TestImportValuesCsv:

    def test_name_value(self):
        b = import_values_csv("Name,Value\nmass,1200\nspeed,25.0\n")
        assert b == {"mass": 1200, "speed": 25.0}

    def test_name_value_unit(self):
        from sysmlpy.usage import ureg
        b = import_values_csv("Name,Value,Unit\nmass,1200,kg\n")
        assert b == {"mass": 1200 * ureg.kg}

    def test_element_attribute_layout(self):
        b = import_values_csv(
            "Element,Attribute,Value\nVehicleSpec::Vehicle,mass,1200\n"
        )
        assert b == {"VehicleSpec::Vehicle::mass": 1200}

    def test_bool_and_string(self):
        b = import_values_csv("Name,Value\nflag,true\nlabel,hello\n")
        assert b["flag"] is True
        assert b["label"] == "hello"

    def test_blank_and_extra_rows(self):
        b = import_values_csv("Name,Value\n\nmass,5\n\nspeed,1\n")
        assert b == {"mass": 5, "speed": 1}

    def test_bad_header_raises(self):
        with pytest.raises(ValueError, match="header row"):
            import_values_csv("Foo,Bar\nbaz,qux\n")

    def test_missing_value_raises(self):
        with pytest.raises(ValueError, match="missing value"):
            import_values_csv("Name,Value\nmass,\n")

    def test_bare_names_work_as_bindings(self):
        m = loads(MODEL)
        bindings = import_values_csv("Name,Value\nmass,1200\n")
        v = evaluate_expression("mass * speed", model=m, bindings=bindings)
        assert v == 1200 * 25.0

    def test_binding_flips_constraint(self):
        m = loads(MODEL)
        report = check_constraints(
            m, bindings=import_values_csv("Name,Value\nmass,6000\n")
        )
        assert not report.failed  # c1 and c3 both pass at mass=6000


# ---------------------------------------------------------------------------
# XLSX import
# ---------------------------------------------------------------------------


class TestImportValuesXlsx:

    @needs_openpyxl
    def test_round_trip_through_xlsx(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Value", "Unit"])
        ws.append(["mass", "1200", "kg"])
        ws.append(["speed", "80"])
        f = tmp_path / "values.xlsx"
        wb.save(f)
        from sysmlpy.usage import ureg
        b = import_values_xlsx(f)
        assert b == {"mass": 1200 * ureg.kg, "speed": 80}

    def test_missing_openpyxl_error(self, tmp_path):
        if _has_openpyxl():
            pytest.skip("openpyxl installed — cannot simulate absence here")
        with pytest.raises(ImportError):
            import_values_xlsx(tmp_path / "none.xlsx")


# ---------------------------------------------------------------------------
# CLI integration
# ---------------------------------------------------------------------------


class TestCli:

    def _model_file(self, tmp_path):
        f = tmp_path / "m.sysml"
        f.write_text(loads(MODEL).dump() + "\n")
        return f

    def test_view_csv_format(self, tmp_path, capsys):
        from sysmlpy.__main__ import main
        assert main(["view", str(self._model_file(tmp_path)),
                     "--view", "tabular", "--format", "csv"]) == 0
        out = capsys.readouterr().out
        assert out.startswith("Name,Type,Kind,Parent")

    def test_eval_set_file_csv(self, tmp_path, capsys):
        f = self._model_file(tmp_path)
        values = tmp_path / "values.csv"
        values.write_text("Name,Value\nmass,6000\n")
        from sysmlpy.__main__ import main
        assert main([
            "eval", str(f), "--constraints", "--set-file", str(values),
        ]) == 0
        assert "2 passed" in capsys.readouterr().out

    def test_eval_set_file_overridden_by_set(self, tmp_path, capsys):
        f = self._model_file(tmp_path)
        values = tmp_path / "values.csv"
        values.write_text("Name,Value\nmass,6000\n")
        from sysmlpy.__main__ import main
        # --set wins over the file: mass=50 fails c1 and c3
        assert main([
            "eval", str(f), "--constraints", "--set-file", str(values),
            "--set", "mass=50",
        ]) == 1

    def test_eval_set_file_missing(self, tmp_path):
        from sysmlpy.__main__ import main
        assert main([
            "eval", str(self._model_file(tmp_path)), "--constraints",
            "--set-file", str(tmp_path / "nope.csv"),
        ]) == 2

    def test_eval_set_file_bad_layout(self, tmp_path):
        f = self._model_file(tmp_path)
        values = tmp_path / "bad.csv"
        values.write_text("Foo,Bar\nbaz,qux\n")
        from sysmlpy.__main__ import main
        assert main([
            "eval", str(f), "--constraints", "--set-file", str(values),
        ]) == 1

    @needs_openpyxl
    def test_xlsx_command(self, tmp_path):
        from sysmlpy.__main__ import main
        out = tmp_path / "model.xlsx"
        assert main([
            "xlsx", str(self._model_file(tmp_path)), "-o", str(out),
        ]) == 0
        import openpyxl
        wb = openpyxl.load_workbook(out)
        assert "Tabular" in wb.sheetnames

    def test_xlsx_missing_output_arg(self, tmp_path, capsys):
        from sysmlpy.__main__ import main
        # -o is required: argparse exits the process with code 2
        with pytest.raises(SystemExit):
            main(["xlsx", str(self._model_file(tmp_path))])

    def test_xlsx_missing_openpyxl(self, tmp_path, monkeypatch):
        if _has_openpyxl():
            pytest.skip("openpyxl installed — cannot simulate absence here")
        from sysmlpy.__main__ import main
        assert main([
            "xlsx", str(self._model_file(tmp_path)), "-o",
            str(tmp_path / "x.xlsx"),
        ]) == 1


# ---------------------------------------------------------------------------
# exports / end-to-end
# ---------------------------------------------------------------------------


def test_public_exports():
    import sysmlpy
    assert sysmlpy.tabular_view_to_csv is tabular_view_to_csv
    assert sysmlpy.import_values_csv is import_values_csv
    assert sysmlpy.import_values_xlsx is import_values_xlsx
    assert sysmlpy.parse_value_literal is parse_value_literal
    assert sysmlpy.write_xlsx is write_xlsx


def test_eval_with_imported_bindings(tmp_path):
    """Full round trip: CSV values file gates model constraints."""
    m = loads(MODEL)
    f = tmp_path / "v.csv"
    f.write_text("Name,Value\nmass,6000\n")
    report = check_constraints(m, bindings=import_values_csv(f.read_text()))
    assert len(report.passed) == 2