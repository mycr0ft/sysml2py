# -*- coding: utf-8 -*-
"""Spreadsheet bridge (v0.66.0 — Adoption Roadmap Goal 7).

Two directions:

**Export** — tabular views to CSV/XLSX workbooks:

    from sysmlpy import loads, tabular_view_csv, write_xlsx
    model = loads(sysml_text)
    csv_text = tabular_view_to_csv(model)             # Name,Type,Kind,...
    write_xlsx(model, "model.xlsx")                   # one sheet per view

**Import** — attribute values from a spreadsheet into evaluator
bindings (what-if runs, constraint gates):

    from sysmlpy import import_values_csv, check_constraints
    bindings = import_values_csv("values.csv")
    report = check_constraints(model, bindings=bindings)

Imported CSV/XLSX files accept either header set:

    Name,Value[,Unit]                 e.g. mass,1200 kg
    Element,Attribute,Value[,Unit]    qualified element + attribute name

Values parse as int → float → boolean (`true`/`false`) → pint unit
string (e.g. ``1200 kg``) → raw string.  Names may be bare
(``mass``) or qualified (``VehicleSpec::Vehicle::mass``).
"""

from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional, Sequence

__all__ = [
    "tabular_view_to_csv", "data_value_tabular_to_csv",
    "relationship_matrix_to_csv", "write_xlsx",
    "import_values_csv", "import_values_xlsx", "parse_value_literal",
]


# ---------------------------------------------------------------------------
# value parsing (shared by import + CLI --set)
# ---------------------------------------------------------------------------


def parse_value_literal(raw: str) -> Any:
    """Parse a spreadsheet/CLI value literal into a Python value.

    Order: bool → int → float → pint quantity → raw string.  Pint
    strings like ``1200 kg`` or ``80 km/h`` become ``Quantity`` objects;
    anything that parses nowhere is kept as a string (e.g. names).
    """
    from sysmlpy.usage import ureg

    text = raw.strip() if isinstance(raw, str) else raw
    low = str(text).strip().lower()
    if low == "true":
        return True
    if low == "false":
        return False
    s = str(text)
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    try:
        return ureg(s)
    except Exception:
        pass
    return s


# ---------------------------------------------------------------------------
# CSV export of the tabular views
# ---------------------------------------------------------------------------


def tabular_view_to_csv(model, focus=None, columns=None) -> str:
    """Tabular View (Name/Type/Kind/Parent/…) as CSV text."""
    from sysmlpy.plantuml import as_tabular_view
    return as_tabular_view(model, focus=focus, output_format="csv",
                           columns=columns)


def data_value_tabular_to_csv(model, focus=None, include_units=True) -> str:
    """Data Value Tabular View (Element/Attribute/Value/Unit/…) as CSV."""
    from sysmlpy.plantuml import as_data_value_tabular_view
    return as_data_value_tabular_view(model, focus=focus,
                                      output_format="csv",
                                      include_units=include_units)


def relationship_matrix_to_csv(model, focus=None, row_type=None,
                               col_type=None, symmetric=True) -> str:
    """Relationship Matrix View as CSV."""
    from sysmlpy.plantuml import as_relationship_matrix_view
    return as_relationship_matrix_view(model, focus=focus,
                                       row_type=row_type, col_type=col_type,
                                       symmetric=symmetric,
                                       output_format="csv")


# ---------------------------------------------------------------------------
# XLSX export (openpyxl optional)
# ---------------------------------------------------------------------------


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ImportError(
            "XLSX export requires the 'openpyxl' package — "
            "install with: pip install 'sysmlpy[xlsx]' (or pip install openpyxl)"
        ) from None
    import openpyxl
    return openpyxl


def write_xlsx(model, path, include=("tabular", "data_value", "matrix"),
               focus=None) -> str:
    """Write tabular views into an Excel workbook (one sheet per view).

    *include* selects sheets: ``"tabular"``, ``"data_value"``,
    ``"matrix"``.  Requires openpyxl (``pip install openpyxl``).
    Returns the resolved output path.
    """
    from pathlib import Path

    openpyxl = _require_openpyxl()

    sheet_sources = {
        "tabular": lambda: tabular_view_to_csv(model, focus=focus),
        "data_value": lambda: data_value_tabular_to_csv(model, focus=focus),
        "matrix": lambda: relationship_matrix_to_csv(model, focus=focus),
    }
    sheet_names = {
        "tabular": "Tabular",
        "data_value": "DataValues",
        "matrix": "Matrix",
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wrote_any = False
    for key in include:
        if key not in sheet_sources:
            raise ValueError(f"Unknown sheet kind: {key!r} — expected one "
                             f"of {sorted(sheet_sources)}")
        rows = list(csv.reader(io.StringIO(sheet_sources[key]())))
        if not rows:
            continue
        ws = wb.create_sheet(sheet_names[key])
        for row in rows:
            ws.append(row)
        # bold header row for readability
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        wrote_any = True

    if not wrote_any:
        # still produce a valid workbook with a note sheet
        ws = wb.create_sheet("Empty")
        ws.append(["(no tabular content in this model)"])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)


# ---------------------------------------------------------------------------
# Value import → evaluator bindings
# ---------------------------------------------------------------------------

_VALUE_HEADERS = ("value", "val")
_NAME_HEADERS = ("name", "element", "qualified name", "qualified_name")
_ATTRIBUTE_HEADERS = ("attribute", "feature", "attr")
_UNIT_HEADERS = ("unit", "units")


def _normalize_header(cell) -> str:
    return str(cell).strip().lower() if cell is not None else ""


def _rows_to_bindings(rows: List[List[Any]]) -> Dict[str, Any]:
    """Turn spreadsheet rows into evaluator ``bindings``.

    Accepted layouts (header row required, case-insensitive):
      Name,Value[,Unit]            — Name may be bare or qualified
      Element,Attribute,Value[,Unit]
    """
    if not rows:
        return {}
    header = [_normalize_header(c) for c in rows[0]]
    col = {}
    for idx, h in enumerate(header):
        if h in _NAME_HEADERS and "name" not in col:
            col["name"] = idx
        elif h in _ATTRIBUTE_HEADERS and "attr" not in col:
            col["attr"] = idx
        elif h in _VALUE_HEADERS and "value" not in col:
            col["value"] = idx
        elif h in _UNIT_HEADERS and "unit" not in col:
            col["unit"] = idx
    if "value" not in col or ("name" not in col and "attr" not in col):
        raise ValueError(
            "Import CSV needs a header row with Name/Value (or "
            "Element/Attribute/Value) columns — "
            f"got header {rows[0]!r}"
        )

    bindings: Dict[str, Any] = {}
    for lineno, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or str(c).strip() == ""
                          for c in row):
            continue  # blank row
        raw_value = row[col["value"]] if col["value"] < len(row) else None
        raw_value = str(raw_value).strip() if raw_value is not None else ""
        if not raw_value:
            raise ValueError(f"Row {lineno}: missing value for "
                             f"{row[:2]!r}")
        if "unit" in col and col["unit"] < len(row) and row[col["unit"]]:
            raw_value = f"{raw_value} {str(row[col['unit']]).strip()}"
        value = parse_value_literal(raw_value)
        if "attr" in col:
            attr = str(row[col["attr"]]).strip() \
                if col["attr"] < len(row) else ""
            base = str(row[col["name"]]).strip() \
                if col["name"] < len(row) else ""
            name = f"{base}::{attr}" if base and attr else (attr or base)
        else:
            name = str(row[col["name"]]).strip()
        bindings[name] = value
    return bindings


def import_values_csv(text: str) -> Dict[str, Any]:
    """Import attribute values from CSV text → evaluator bindings dict.

    See the module docstring for the accepted layouts.
    """
    rows = list(csv.reader(io.StringIO(text)))
    return _rows_to_bindings(rows)


def import_values_xlsx(path) -> Dict[str, Any]:
    """Import attribute values from the first sheet of an XLSX workbook."""
    openpyxl = _require_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    return _rows_to_bindings(rows)